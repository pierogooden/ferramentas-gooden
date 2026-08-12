import base64
import io
import json
import sys
from pathlib import Path

import fitz  # pymupdf
import streamlit as st
from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from utils.design import inject_css, render_header

st.set_page_config(
    page_title="Canivete Gooden",
    page_icon="🚌",
    layout="centered",
)

inject_css()
render_header("Listas de Passageiros", "Extração automática", page_key="listas")

MEDIA_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


def pdf_para_imagens(pdf_bytes: bytes) -> list[bytes]:
    """Converte cada página do PDF em PNG (bytes) para envio à API."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    paginas = []
    for pagina in doc:
        # Renderiza em alta resolução (2x para melhor OCR)
        mat = fitz.Matrix(2.0, 2.0)
        pix = pagina.get_pixmap(matrix=mat)
        paginas.append(pix.tobytes("png"))
    doc.close()
    return paginas


# ── Funções ──────────────────────────────────────────────────────────────────
def extrair_passageiros(client: Groq, imagem_bytes: bytes, media_type: str) -> list[dict]:
    imagem_b64 = base64.standard_b64encode(imagem_bytes).decode("utf-8")
    prompt = """Analise esta imagem de uma lista de passageiros e extraia TODOS os dados visíveis.

Para cada passageiro encontrado, retorne um objeto JSON com:
- "nome": nome completo do passageiro
- "documento": SOMENTE os dígitos e pontuação do número (ex: "12.345.678-9" ou "123.456.789-00"). NÃO inclua o tipo do documento (não escreva "RG", "CPF", "Doc", "Passaporte" etc.). Se houver mais de um número, retorne apenas o principal.
- "observacao": qualquer informação adicional relevante (opcional, deixe vazio se não houver)

Se um campo não estiver visível ou legível, use null.

Responda SOMENTE com um array JSON válido, sem texto adicional, sem markdown, sem explicações.
Exemplo: [{"nome": "João Silva", "documento": "12.345.678-9", "observacao": ""}, ...]

Se a imagem não contiver lista de passageiros ou não for legível, retorne: []"""

    response = client.chat.completions.create(
        model="qwen/qwen3.6-27b",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{imagem_b64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        max_tokens=4096,
        temperature=0,
        reasoning_effort="none",
    )
    texto = response.choices[0].message.content.strip()

    import re as _re
    texto = _re.sub(r"<think>.*?</think>", "", texto, flags=_re.DOTALL).strip()

    if texto.startswith("```"):
        linhas = texto.split("\n")
        texto = "\n".join(linhas[1:-1] if linhas[-1].strip() == "```" else linhas[1:])

    try:
        resultado = json.loads(texto)
        if not isinstance(resultado, list):
            return []
        # Remove rótulos de tipo de documento que possam ter vindo junto
        import re
        prefixos = re.compile(
            r"^\s*(rg|cpf|doc\.?|documento|passaporte|cnh|rne|ctps|pis|nit"
            r"|id|identidade|n[°º]?\.?)\s*[:\-]?\s*",
            re.IGNORECASE,
        )
        for p in resultado:
            if p.get("documento"):
                p["documento"] = prefixos.sub("", str(p["documento"])).strip()
        return resultado
    except json.JSONDecodeError:
        return []


def gerar_xlsx(passageiros: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Passageiros"

    cor_header = "020066"
    cor_par    = "F4F5FF"
    f_header   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fill_header = PatternFill(fill_type="solid", fgColor=cor_header)
    fill_par    = PatternFill(fill_type="solid", fgColor=cor_par)

    cabecalhos = ["✓", "#", "Nome Completo", "Documento", "Observação"]
    larguras   = [5,   6,   42,              20,           28]

    for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), start=1):
        cel = ws.cell(row=1, column=col, value=cab)
        cel.font = f_header
        cel.fill = fill_header
        cel.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cel.column_letter].width = larg
    ws.row_dimensions[1].height = 22

    for i, p in enumerate(passageiros, start=1):
        linha = i + 1
        ws.cell(row=linha, column=1, value="☐").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=linha, column=1).font = Font(name="Calibri", size=13)
        for col, val in enumerate([i, p.get("nome") or "", p.get("documento") or "", p.get("observacao") or ""], start=2):
            cel = ws.cell(row=linha, column=col, value=val)
            cel.alignment = Alignment(vertical="center", wrap_text=True)
        if linha % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=linha, column=col).fill = fill_par
        ws.row_dimensions[linha].height = 18

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"B1:{ws.cell(row=1, column=5).coordinate}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── API ───────────────────────────────────────────────────────────────────────
try:
    api_key = st.secrets["GROQ_API_KEY"]
except Exception:
    import os
    api_key = os.getenv("GROQ_API_KEY", "")

if not api_key:
    st.error("⚠️ Chave GROQ_API_KEY não configurada nos Secrets.")
    st.stop()

client = Groq(api_key=api_key)

# ── Upload ────────────────────────────────────────────────────────────────────
st.markdown('<div class="g-section-label">Imagens das listas</div>', unsafe_allow_html=True)

arquivos = st.file_uploader(
    "Selecione ou arraste os arquivos",
    type=["jpg", "jpeg", "png", "webp", "pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if not arquivos:
    st.markdown("""
    <div class="g-empty">
        <div class="g-empty-icon">📋</div>
        <div class="g-empty-title">Nenhum arquivo carregado</div>
        <div class="g-empty-sub">Suporta JPG, PNG, WEBP e PDF · Uma planilha gerada por arquivo</div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Resultados ────────────────────────────────────────────────────────────────
st.markdown(f'<div class="g-section-label" style="margin-top:28px">{len(arquivos)} arquivo(s) carregado(s)</div>', unsafe_allow_html=True)

for arquivo in arquivos:
    sufixo    = Path(arquivo.name).suffix.lower()
    nome_base = Path(arquivo.name).stem
    eh_pdf    = sufixo == ".pdf"
    icone     = "📑" if eh_pdf else "📄"

    with st.expander(f"{icone}  {arquivo.name}", expanded=True):
        arquivo_bytes = arquivo.read()

        # ── PDF: converte páginas em imagens ──
        if eh_pdf:
            with st.spinner("Convertendo páginas do PDF..."):
                paginas = pdf_para_imagens(arquivo_bytes)

            n_pags = len(paginas)
            st.markdown(
                f'<div style="font-family:Geologica,sans-serif;font-size:0.78rem;'
                f'color:#ACB0F8;margin-bottom:12px">'
                f'PDF com {n_pags} página(s)</div>',
                unsafe_allow_html=True,
            )

            todos = []
            for idx, pag_bytes in enumerate(paginas, start=1):
                with st.spinner(f"Analisando página {idx}/{n_pags}..."):
                    resultado = extrair_passageiros(client, pag_bytes, "image/png")
                    todos.extend(resultado)

            passageiros = todos

        # ── Imagem normal ──
        else:
            col_img, col_info = st.columns([1, 2])
            with col_img:
                st.image(arquivo_bytes, use_container_width=True)
            with col_info:
                with st.spinner("Analisando imagem..."):
                    passageiros = extrair_passageiros(client, arquivo_bytes, MEDIA_TYPES.get(sufixo, "image/jpeg"))

        # ── Resultado ──
        if not passageiros:
            st.warning("Nenhum passageiro encontrado neste arquivo.")
        else:
            n = len(passageiros)

            if eh_pdf:
                st.markdown(f"""
                <div style="margin-bottom:12px">
                    <span style="font-family:'Geologica',sans-serif;font-weight:700;
                                 font-size:1.6rem;color:#020066;">{n}</span>
                    <span style="font-family:'Abhaya Libre',serif;color:#8386C8;
                                 font-size:0.9rem;margin-left:6px;">passageiro(s) em {len(paginas)} página(s)</span>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="margin-bottom:12px">
                    <span style="font-family:'Geologica',sans-serif;font-weight:700;
                                 font-size:1.6rem;color:#020066;">{n}</span>
                    <span style="font-family:'Abhaya Libre',serif;color:#8386C8;
                                 font-size:0.9rem;margin-left:6px;">passageiro(s) identificado(s)</span>
                </div>
                """, unsafe_allow_html=True)

            st.dataframe(
                [{
                    "✓": "☐",
                    "#": i + 1,
                    "Nome": p.get("nome") or "—",
                    "Documento": p.get("documento") or "—",
                    "Obs.": p.get("observacao") or "",
                } for i, p in enumerate(passageiros)],
                use_container_width=True,
                hide_index=True,
            )

            st.download_button(
                label=f"⬇  Baixar  {nome_base}.xlsx",
                data=gerar_xlsx(passageiros),
                file_name=f"{nome_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown('<div class="g-footer">Gooden · Conduzindo tranquilidade</div>', unsafe_allow_html=True)
