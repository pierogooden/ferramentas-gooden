"""
Utilitários para escrita de dados de OS nas planilhas Excel.
"""
import io
import re
from collections import defaultdict
from datetime import date, datetime, time
from typing import Optional

from openpyxl import load_workbook


# ── Conversões ────────────────────────────────────────────────────────────────

def parse_prefixo(carro) -> int:
    """'0001283' → 1283"""
    return int(str(carro).strip().lstrip("0") or "0")


def time_to_fraction(h: int, m: int) -> float:
    """Converte hora e minuto para fração decimal do dia (formato Excel)."""
    return (h * 60 + m) / 1440.0


def parse_hora_str(s) -> Optional[float]:
    """'03:20' ou '3:20' → fração decimal do dia para Excel. Aceita '26:50' → 02:50."""
    if s is None:
        return None
    txt = str(s).strip().replace(".", ":")
    m = re.match(r"(\d{1,2}):(\d{2})", txt)
    if not m:
        return None
    h = int(m.group(1)) % 24
    mn = int(m.group(2))
    return time_to_fraction(h, mn)


def parse_data_str(s) -> Optional[date]:
    """'27/07/2026' ou '27/07/26' → date."""
    if not s:
        return None
    try:
        parts = str(s).strip().split("/")
        d, mo, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
        return date(y, mo, d)
    except Exception:
        return None


def nome_curto(nome: str) -> str:
    """'CARLOS LUIZ DA SILVA' → 'Carlos Luiz'"""
    partes = str(nome).strip().title().split()
    return " ".join(partes[:2]) if len(partes) >= 2 else nome.title()


def salvar_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_data(d) -> Optional[date]:
    if isinstance(d, (date, datetime)):
        return d.date() if isinstance(d, datetime) else d
    return parse_data_str(d)


def get_hora(h) -> Optional[float]:
    if isinstance(h, time):
        return time_to_fraction(h.hour, h.minute)
    if isinstance(h, float):
        return h
    return parse_hora_str(h)


def get_valor_viagem(comissao) -> Optional[float]:
    """Comissão é 10% do valor. Valor = comissao * 10."""
    try:
        v = float(comissao)
        return round(v * 10, 2) if v else None
    except (TypeError, ValueError):
        return None


# ── Localizadores de aba ──────────────────────────────────────────────────────

_KM_NOMES = {
    1: ["Janeiro 2026", "Janeiro 2026 "],
    2: ["Fevereiro 2026", "Fevereiro 2026 "],
    3: ["Março 2026", "Março 2026 ", "Marco 2026"],
    4: ["Abril 2026", "Abril 2026 "],
    5: ["Maio 2026", "Maio 2026 "],
    6: ["Junho 2026", "Junho 2026 "],
    7: ["Julho 2026", "Julho 2026 "],
    8: ["Agosto 2026", "Agosto 2026 "],
    9: ["Setembro 2026", "Setembro 2026 "],
    10: ["Outubro 2026", "Outubro 2026 "],
    11: ["Novembro 2026", "Novembro 2026 "],
    12: ["Dezembro 2026", "Dezembro 2026 "],
}

_ANALISE_NOMES = {
    1: "01. Jan", 2: "02. Fev", 3: "03. Mar", 4: "04. Abr",
    5: "05. Mai", 6: "06. Jun", 7: "07. Jul", 8: "08. Ago",
    9: "09. Set", 10: "10. Out", 11: "11. Nov", 12: "12. Dez",
}


def find_km_tab(wb, mes: int) -> Optional[str]:
    sheets = wb.sheetnames
    for c in _KM_NOMES.get(mes, []):
        if c in sheets:
            return c
    # busca parcial (strip)
    for name in sheets:
        for c in _KM_NOMES.get(mes, []):
            if c.strip().lower() == name.strip().lower():
                return name
    return None


def find_analise_tab(wb, mes: int) -> Optional[str]:
    candidato = _ANALISE_NOMES.get(mes)
    if candidato and candidato in wb.sheetnames:
        return candidato
    return None


def find_comissao_tab(wb, data_os: date) -> Optional[str]:
    """Encontra a aba semanal que contém a data da OS."""
    for name in wb.sheetnames:
        m = re.match(r"(\d{2})\.(\d{2})\s*[–\-a]\s*(\d{2})\.(\d{2})", name.strip())
        if not m:
            continue
        d1, m1, d2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        year = data_os.year
        start = date(year, m1, d1)
        end_year = year + 1 if m2 < m1 else year
        try:
            end = date(end_year, m2, d2)
        except ValueError:
            continue
        if start <= data_os <= end:
            return name
    return None


# ── Próxima linha vazia ───────────────────────────────────────────────────────

def _next_row(ws, start: int, col_check: int) -> int:
    for r in range(start, ws.max_row + 200):
        if ws.cell(r, col_check).value is None:
            return r
    return ws.max_row + 1


def next_row_km(ws) -> int:
    return _next_row(ws, 2, 1)


def next_row_analise(ws) -> int:
    return _next_row(ws, 6, 3)  # col C = data_ida


def next_row_comissao(ws) -> int:
    return _next_row(ws, 3, 1)


# ── Formatos numéricos ─────────────────────────────────────────────────────────

FMT_HORA = "HH:MM"
FMT_DATA = "DD/MM/YYYY"
FMT_KM   = "#,##0"
FMT_MOEDA = 'R$ #,##0.00'
FMT_PCT = "0.00%"


def _fmt(ws, row, col, num_format):
    ws.cell(row, col).number_format = num_format


# ── Escritores de linha ────────────────────────────────────────────────────────

def _write_km_row(ws, r: int, d: dict):
    """Escreve uma linha na planilha Km Turismo."""
    data_os = get_data(d.get("data_saida"))

    ws.cell(r, 1).value = d.get("num_os")
    ws.cell(r, 2).value = data_os
    _fmt(ws, r, 2, FMT_DATA)

    ws.cell(r, 3).value = nome_curto(d.get("motorista_nome", ""))
    ws.cell(r, 4).value = d.get("destino", "")
    ws.cell(r, 5).value = (
        parse_prefixo(d.get("carro", d.get("prefixo", 0)))
        if not isinstance(d.get("prefixo"), int)
        else d["prefixo"]
    )

    # Odômetros + horários (cols F=6 a U=21)
    campos = [
        ("saida_garagem_km",     "saida_garagem_hora"),
        ("chegada_cliente_km",   "chegada_cliente_hora"),
        ("saida_cliente_km",     "saida_cliente_hora"),
        ("chegada_destino_km",   "chegada_destino_hora"),
        ("saida_destino_km",     "saida_destino_hora"),
        ("chegada_origem_km",    "chegada_origem_hora"),
        # col R,S = repetição da origem (sem saída distinta)
        ("chegada_origem_km",    "chegada_origem_hora"),
        ("chegada_garagem_km",   "chegada_garagem_hora"),
    ]
    col = 6
    for km_key, hr_key in campos:
        ws.cell(r, col).value = d.get(km_key)
        _fmt(ws, r, col, FMT_KM)
        col += 1
        h = get_hora(d.get(hr_key))
        ws.cell(r, col).value = h
        if h is not None:
            _fmt(ws, r, col, FMT_HORA)
        col += 1

    # col 22=V: KM PRODUTIVO
    ws.cell(r, 22).value = f"=(L{r}-J{r})+(P{r}-N{r})"
    # col 23=W: KM OCIOSO
    ws.cell(r, 23).value = f"=X{r}-V{r}"
    # col 24=X: KM TOTAL
    ws.cell(r, 24).value = f"=T{r}-F{r}"
    # col 25=Y: KM CONTRATADO (manual, blank)
    # col 26=Z: HORARIO TOTAL
    ws.cell(r, 26).value = f"=MOD(U{r}-G{r},1)"
    _fmt(ws, r, 26, FMT_HORA)
    # col 27=AA: Valor da Viagem
    ws.cell(r, 27).value = d.get("valor_viagem")
    if d.get("valor_viagem"):
        _fmt(ws, r, 27, FMT_MOEDA)


def _write_analise_row(ws, r: int, d: dict, seq: int):
    """Escreve uma linha na planilha Análise de Resultados."""
    data_saida  = get_data(d.get("data_saida"))
    data_retorno = get_data(d.get("data_retorno")) or data_saida

    ws.cell(r, 1).value = None
    ws.cell(r, 2).value = f'=IF(C{r}="","",{seq})'
    ws.cell(r, 3).value = data_saida
    _fmt(ws, r, 3, FMT_DATA)
    ws.cell(r, 4).value = data_retorno
    _fmt(ws, r, 4, FMT_DATA)
    ws.cell(r, 5).value = d.get("num_os")
    ws.cell(r, 6).value = d.get("num_os")   # Contrato = OS
    ws.cell(r, 7).value = None               # CPF/CNPJ - manual
    ws.cell(r, 8).value = d.get("contratante", "")
    ws.cell(r, 9).value = None               # Nota fiscal - manual
    ws.cell(r, 10).value = None              # Centro de Custos - manual
    ws.cell(r, 11).value = d.get("valor_viagem")  # Receita
    if d.get("valor_viagem"):
        _fmt(ws, r, 11, FMT_MOEDA)
    pref = (parse_prefixo(d.get("carro", d.get("prefixo", 0)))
            if not isinstance(d.get("prefixo"), int) else d["prefixo"])
    ws.cell(r, 12).value = pref             # Prefixo
    ws.cell(r, 13).value = None             # Km total - manual
    # Alíquotas
    ws.cell(r, 14).value = 0.0065           # PIS
    ws.cell(r, 15).value = 0.03             # Cofins
    ws.cell(r, 16).value = 0               # ISS
    ws.cell(r, 17).value = 0.12             # ICMS
    # Fórmulas
    ws.cell(r, 18).value = f'=IF(K{r}="","",SUM(N{r}:Q{r})*K{r})'   # Imposto Pago
    ws.cell(r, 19).value = 0               # Devolução
    ws.cell(r, 20).value = None             # Pedágio - manual
    ws.cell(r, 21).value = None             # Estacionamento - manual
    ws.cell(r, 22).value = None             # Outros custos - manual
    ws.cell(r, 23).value = None             # Custo variável - manual
    ws.cell(r, 24).value = f'=IF(K{r}="","",K{r}*10%)'    # Comissão motorista
    ws.cell(r, 25).value = f'=IF(K{r}="","",K{r}*2%)'     # Comissão time vendas
    ws.cell(r, 26).value = f'=IF(SUM(R{r}:Y{r})=0,"",SUM(R{r}:Y{r}))'  # Custo total
    ws.cell(r, 27).value = f'=IF(K{r}="","",K{r}-Z{r})'   # Gross Margin R$
    ws.cell(r, 28).value = f'=IF(K{r}="","",AA{r}/K{r})'  # Gross Margin %
    _fmt(ws, r, 28, FMT_PCT)


def _write_comissao_row(ws, r: int, d: dict, formula_pagar: Optional[str]):
    """Escreve uma linha na planilha Comissão dos Motoristas."""
    data_os = get_data(d.get("data_saida"))

    ws.cell(r, 1).value = d.get("motorista_nome", "")
    ws.cell(r, 2).value = None              # Filial - manual
    ws.cell(r, 3).value = int(str(d.get("motorista_matricula", "0")).lstrip("0") or "0")
    ws.cell(r, 4).value = d.get("num_os")
    ws.cell(r, 5).value = None             # Centro de Custo - manual
    ws.cell(r, 6).value = data_os
    if data_os:
        _fmt(ws, r, 6, FMT_DATA)
    ws.cell(r, 7).value = d.get("valor_viagem")
    if d.get("valor_viagem"):
        _fmt(ws, r, 7, FMT_MOEDA)
    ws.cell(r, 8).value = f"=G{r}*10%"     # Comissão 10%
    _fmt(ws, r, 8, FMT_MOEDA)
    ws.cell(r, 9).value = formula_pagar    # Comissões a pagar


# ── Agrupamento para Comissão ─────────────────────────────────────────────────

def _agrupar_comissao(dados_list: list) -> list:
    """
    Retorna lista de (d, formula_pagar) agrupando por matrícula.
    O primeiro da mesma matrícula recebe a fórmula de soma.
    """
    # Agrupa mantendo a ordem original
    grupos: dict[str, list] = {}
    for d in dados_list:
        mat = str(d.get("motorista_matricula", ""))
        grupos.setdefault(mat, []).append(d)

    result = []
    base_row = [None]  # será atualizado durante escrita

    # Constrói sequência de (d, formula_pagar_placeholder)
    for mat, items in grupos.items():
        for i, d in enumerate(items):
            result.append((d, i, len(items)))

    return result


# ── Escritores principais ──────────────────────────────────────────────────────

def escrever_km(wb, dados_list: list, aba: str = None):
    """Escreve dados na aba correta do Km Turismo."""
    for d in dados_list:
        data_os = get_data(d.get("data_saida"))
        tab = aba or (find_km_tab(wb, data_os.month) if data_os else None)
        if not tab or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        r = next_row_km(ws)
        _write_km_row(ws, r, d)


def escrever_analise(wb, dados_list: list, aba: str = None):
    """Escreve dados na aba correta da Análise de Resultados."""
    seq_por_aba: dict[str, int] = defaultdict(int)
    for d in dados_list:
        data_os = get_data(d.get("data_saida"))
        tab = aba or (find_analise_tab(wb, data_os.month) if data_os else None)
        if not tab or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        # Calcula seq: quantas linhas já existem nessa aba
        if seq_por_aba[tab] == 0:
            for r in range(6, ws.max_row + 1):
                if ws.cell(r, 3).value is not None:
                    seq_por_aba[tab] += 1
        seq_por_aba[tab] += 1
        r = next_row_analise(ws)
        _write_analise_row(ws, r, d, seq_por_aba[tab])


def escrever_comissao(wb, dados_list: list, aba: str = None):
    """Escreve dados na aba correta da Comissão dos Motoristas."""
    # Agrupa por aba
    por_aba: dict[str, list] = defaultdict(list)
    for d in dados_list:
        data_os = get_data(d.get("data_saida"))
        tab = aba or (find_comissao_tab(wb, data_os) if data_os else None)
        if tab and tab in wb.sheetnames:
            por_aba[tab].append(d)

    for tab, items in por_aba.items():
        ws = wb[tab]
        base_r = next_row_comissao(ws)

        # Reagrupa por matrícula mantendo ordem de aparição
        grupos: dict[str, list] = {}
        for d in items:
            mat = str(d.get("motorista_matricula", ""))
            grupos.setdefault(mat, []).append(d)

        cur = base_r
        row_map: list[tuple] = []  # (d, row, idx_no_grupo, total_no_grupo)
        for mat, grp in grupos.items():
            for i, d in enumerate(grp):
                row_map.append((d, cur, i, len(grp)))
                cur += 1

        for d, r, i, total in row_map:
            if total == 1:
                formula = f"=H{r}"
            elif i == 0:
                formula = f"=SUM(H{r}:H{r + total - 1})"
            else:
                formula = None
            _write_comissao_row(ws, r, d, formula)

        # Linha de total
        last = cur - 1
        tot_r = cur
        ws.cell(tot_r, 8).value = "Total "
        ws.cell(tot_r, 9).value = f"=SUM(I{base_r}:I{last})"
        _fmt(ws, tot_r, 9, FMT_MOEDA)


# ── Arquivo único para download ───────────────────────────────────────────────

_HDRS_KM = [
    "FICHA", "DATA", "Motorista", "Destino", "PREFIXO",
    "Km garagem ou residência", "Horário",
    "Km chegada no Cliente", "Horário",
    "Km Cliente", "Horário",
    "Km Chegada no Destino", "Horário",
    "Km Saída Destino", "Horário",
    "Km chegada no Cliente", "Horário",
    "Km Cliente", "Horário",
    "Km Chegada garagem ou residência", "Horário",
    "KM PRODUTIVO", "KM OCIOSO", "KM TOTAL", "KM CONTRATADO",
    "HORARIO TOTAL", "Valor da Viagem",
]

_HDRS_ANALISE = [
    None, "#", "Data ida", "Data volta", "OS", "Contrato",
    "CPF/CNPJ Contratante", "Nome/Razão Social Contratante",
    "Nota fiscal n°", "Centro de Custos (CC)", "Receita",
    "Prefixo do veículo", "Km total", "PIS", "Cofins", "ISS",
    "ICMS", "Imposto Pago", "Devolução para cliente",
    "Pedágio", "Estacionamento", "Outros custos",
    "Custo variável", "Comissão motorista", "Comissão time de vendas",
    "Custo total", "Gross Margin R$", "Gross Margin %",
]

_HDRS_COMISSAO = [
    "Nome", "Filial/Matriz", "Matrícula", "N° do Contrato",
    "Centro de Custo", "Data", "Viagem $", "Comissão 10%", "Comissões a pagar",
]


def gerar_download(dados_list: list) -> bytes:
    """
    Gera um workbook Excel com 3 abas (Km Turismo, Análise, Comissão)
    preenchidas com os dados das OS fornecidas.
    Retorna os bytes prontos para download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # remove aba vazia padrão

    cor_hdr = "020066"
    f_hdr = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fill_hdr = PatternFill(fill_type="solid", fgColor=cor_hdr)

    def _hdr_row(ws, row, colunas):
        for c, txt in enumerate(colunas, 1):
            cel = ws.cell(row, c, value=txt)
            if txt:
                cel.font = f_hdr
                cel.fill = fill_hdr
                cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Aba 1: Km Turismo ──────────────────────────────────────────────────────
    ws_km = wb.create_sheet("Km Turismo")
    _hdr_row(ws_km, 1, _HDRS_KM)
    ws_km.row_dimensions[1].height = 30
    for d in dados_list:
        _write_km_row(ws_km, next_row_km(ws_km), d)

    # ── Aba 2: Análise de Resultados ──────────────────────────────────────────
    ws_an = wb.create_sheet("Análise de Resultados")
    ws_an.cell(2, 4).value = "Planilha de Análise de Resultado | Fretamento Eventual"
    ws_an.cell(2, 4).font = Font(bold=True, size=12)
    # Detecta mês dos dados
    datas = [get_data(d.get("data_saida")) for d in dados_list if d.get("data_saida")]
    if datas:
        meses_pt = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        mes_n = datas[0].month
        ano = datas[0].year
        ws_an.cell(3, 4).value = f"Mês referência: {meses_pt[mes_n]} de {ano}"
    _hdr_row(ws_an, 5, _HDRS_ANALISE)
    ws_an.row_dimensions[5].height = 30
    seq = 0
    for d in dados_list:
        r = next_row_analise(ws_an)
        seq += 1
        _write_analise_row(ws_an, r, d, seq)

    # ── Aba 3: Comissão dos Motoristas ────────────────────────────────────────
    ws_co = wb.create_sheet("Comissão dos Motoristas")
    ws_co.cell(1, 1).value = "COMISSÕES MOTORISTAS"
    ws_co.cell(1, 1).font = Font(bold=True, size=11)
    _hdr_row(ws_co, 2, _HDRS_COMISSAO)
    ws_co.row_dimensions[2].height = 22

    # Agrupa por matrícula
    grupos: dict[str, list] = {}
    for d in dados_list:
        mat = str(d.get("motorista_matricula", ""))
        grupos.setdefault(mat, []).append(d)

    cur = 3
    row_map: list[tuple] = []
    for mat, grp in grupos.items():
        n = len(grp)
        for i, d in enumerate(grp):
            row_map.append((d, cur, i, n))
            cur += 1

    for d, r, i, total in row_map:
        if i == 0 and total == 1:
            formula_pagar = f"=H{r}"
        elif i == 0:
            formula_pagar = f"=SUM(H{r}:H{r + total - 1})"
        else:
            formula_pagar = None
        _write_comissao_row(ws_co, r, d, formula_pagar)

    last = cur - 1
    ws_co.cell(cur, 8).value = "Total "
    ws_co.cell(cur, 9).value = f"=SUM(I3:I{last})"
    _fmt(ws_co, cur, 9, FMT_MOEDA)

    return salvar_bytes(wb)


# ── Criadores de aba TESTE ────────────────────────────────────────────────────

def _ref_km(wb) -> Optional[str]:
    for mes in range(12, 0, -1):
        t = find_km_tab(wb, mes)
        if t:
            return t
    return None


def _ref_analise(wb) -> Optional[str]:
    for mes in range(12, 0, -1):
        t = find_analise_tab(wb, mes)
        if t:
            return t
    return None


def criar_teste_km(wb, dados_list: list) -> str:
    if "TESTE" in wb.sheetnames:
        del wb["TESTE"]

    ref = _ref_km(wb)
    if ref:
        ws = wb.copy_worksheet(wb[ref])
        ws.title = "TESTE"
        # Limpa linhas de dados (mantém cabeçalho row 1)
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None
    else:
        ws = wb.create_sheet("TESTE")
        hdrs = [
            "FICHA", "DATA", "Motorista", "Destino", "PREFIXO",
            "Km garagem", "Horário", "Km chegada Cliente", "Horário",
            "Km Cliente", "Horário", "Km Chegada Destino", "Horário",
            "Km Saída Destino", "Horário", "Km chegada Origem", "Horário",
            "Km Saída Origem", "Horário", "Km Chegada Garagem", "Horário",
            "KM PRODUTIVO", "KM OCIOSO", "KM TOTAL", "KM CONTRATADO",
            "HORARIO TOTAL", "Valor da Viagem",
        ]
        for c, h in enumerate(hdrs, 1):
            ws.cell(1, c).value = h

    for d in dados_list:
        _write_km_row(ws, next_row_km(ws), d)
    return "TESTE"


def criar_teste_analise(wb, dados_list: list) -> str:
    if "TESTE" in wb.sheetnames:
        del wb["TESTE"]

    ref = _ref_analise(wb)
    if ref:
        ws = wb.copy_worksheet(wb[ref])
        ws.title = "TESTE"
        for r in range(6, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None
        # Atualiza título do mês
        for r in range(1, 6):
            for c in range(1, 10):
                v = ws.cell(r, c).value
                if v and "referência" in str(v).lower():
                    ws.cell(r, c).value = "Mês referência: Julho de 2026"
    else:
        ws = wb.create_sheet("TESTE")
        ws.cell(2, 4).value = "Planilha de Análise de Resultado | Fretamento Eventual"
        ws.cell(3, 4).value = "Mês referência: Julho de 2026"
        hdrs = [
            None, "#", "Data ida", "Data volta", "OS", "Contrato",
            "CPF/CNPJ Contratante", "Nome/Razão Social Contratante",
            "Nota fiscal n°", "Centro de Custos (CC)", "Receita",
            "Prefixo do veículo", "Km total", "PIS", "Cofins", "ISS",
            "ICMS", "Imposto Pago", "Devolução para cliente",
            "Pedágio", "Estacionamento", "Outros custos",
            "Custo variável", "Comissão motorista", "Comissão time de vendas",
            "Custo total", "Gross Margin R$", "Gross Margin %",
        ]
        for c, h in enumerate(hdrs, 1):
            ws.cell(5, c).value = h

    seq = 0
    for d in dados_list:
        r = next_row_analise(ws)
        seq += 1
        _write_analise_row(ws, r, d, seq)
    return "TESTE"


def criar_teste_comissao(wb, dados_list: list) -> str:
    if "TESTE" in wb.sheetnames:
        del wb["TESTE"]

    # Cria do zero (evita problemas de copy_worksheet com células protegidas)
    ws = wb.create_sheet("TESTE")
    ws.cell(1, 1).value = "COMISSÕES MOTORISTAS"
    hdrs = [
        "Nome", "Filial/Matriz", "Matrícula", "N° do Contrato",
        "Centro de Custo", "Data", "Viagem $", "Comissão 10%", "Comissões a pagar",
    ]
    for c, h in enumerate(hdrs, 1):
        ws.cell(2, c).value = h

    if not dados_list:
        return "TESTE"

    # Agrupa por matrícula mantendo ordem de aparição
    grupos: dict[str, list] = {}
    for d in dados_list:
        mat = str(d.get("motorista_matricula", ""))
        grupos.setdefault(mat, []).append(d)

    # Constrói mapa de linhas com posição e contexto do grupo
    cur = 3
    row_map: list[tuple] = []
    for mat, grp in grupos.items():
        n = len(grp)
        for i, d in enumerate(grp):
            row_map.append((d, cur, i, n))
            cur += 1

    # Escreve cada linha
    for d, r, i, total in row_map:
        if i == 0 and total == 1:
            formula_pagar = f"=H{r}"
        elif i == 0:
            formula_pagar = f"=SUM(H{r}:H{r + total - 1})"
        else:
            formula_pagar = None
        _write_comissao_row(ws, r, d, formula_pagar)

    # Linha de total
    last = cur - 1
    ws.cell(cur, 8).value = "Total "
    ws.cell(cur, 9).value = f"=SUM(I3:I{last})"
    _fmt(ws, cur, 9, FMT_MOEDA)
    return "TESTE"
