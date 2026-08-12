"""
Script único para criar abas TESTE nas 3 planilhas com os dados das 5 OS de exemplo.
Execute: python criar_teste.py
"""
import sys
from pathlib import Path
from datetime import date, time

sys.path.insert(0, str(Path(__file__).parent))
from utils.planilhas_os import (
    time_to_fraction,
    criar_teste_km, criar_teste_analise, criar_teste_comissao,
)
from openpyxl import load_workbook

# ── Caminhos ──────────────────────────────────────────────────────────────────
ARQUIVOS = {
    "km":      Path(r"C:\Users\Dell\Downloads\Km Turismo - 2026.xlsx"),
    "analise": Path(r"C:\Users\Dell\Desktop\Gooden\Fretamento Eventual - Planilha de análise de resultados.xlsx"),
    "comissao": Path(r"C:\Users\Dell\Downloads\Comissão dos Motoristas – 2026.xlsx"),
}

# ── Dados das 5 OS extraídos das fotos ───────────────────────────────────────
def hr(h, m): return time_to_fraction(h, m)

OS_EXEMPLOS = [
    {
        # OS 8086 · 27/07/26 · Carlos Luiz · São José dos Campos
        "num_os":             "8086",
        "contratante":        "FUNDACAO EDUCACIONAL INACIANA PADRE SABOIA",
        "motorista_matricula": 2404,
        "motorista_nome":     "Carlos Luiz da Silva",
        "carro":              "0001283",
        "prefixo":            1283,
        "destino":            "SAO JOSE DOS CAMPOS",
        "data_saida":         date(2026, 7, 27),
        "data_retorno":       date(2026, 7, 27),
        "hora_saida":         "03:45",
        "hora_retorno":       "18:00",
        "saida_garagem_km":   559383, "saida_garagem_hora":   hr(3, 20),
        "chegada_cliente_km": 559395, "chegada_cliente_hora": hr(3, 40),
        "saida_cliente_km":   559395, "saida_cliente_hora":   hr(6, 10),
        "chegada_destino_km": 559527, "chegada_destino_hora": hr(9, 30),
        "saida_destino_km":   559527, "saida_destino_hora":   hr(18, 35),
        "chegada_origem_km":  559659, "chegada_origem_hora":  hr(20, 50),
        "chegada_garagem_km": 559672, "chegada_garagem_hora": hr(21, 40),
        "comissao":  286.19,
        "valor_viagem": 2861.90,
    },
    {
        # OS 8130 · 31/07/26 · Ananias · Aparecida do Norte
        "num_os":             "8130",
        "contratante":        "PROVINCIA AGOSTINIANA DO BRASIL",
        "motorista_matricula": 3091,
        "motorista_nome":     "Ananias Martins de Oliveira",
        "carro":              "0000489",
        "prefixo":            489,
        "destino":            "APARECIDA DO NORTE",
        "data_saida":         date(2026, 7, 31),
        "data_retorno":       date(2026, 7, 31),
        "hora_saida":         "04:00",
        "hora_retorno":       "16:00",
        "saida_garagem_km":   29022,  "saida_garagem_hora":   hr(2, 25),
        "chegada_cliente_km": 29055,  "chegada_cliente_hora": hr(3, 40),
        "saida_cliente_km":   29055,  "saida_cliente_hora":   hr(4, 35),
        "chegada_destino_km": 29283,  "chegada_destino_hora": hr(8, 22),
        "saida_destino_km":   29283,  "saida_destino_hora":   hr(16, 18),
        "chegada_origem_km":  29499,  "chegada_origem_hora":  hr(21, 10),
        "chegada_garagem_km": 29530,  "chegada_garagem_hora": hr(22, 45),
        "comissao":  358.60,
        "valor_viagem": 3586.00,
    },
    {
        # OS 8243 · 30/07/26 · Carlos Luiz · São Paulo Liberdade
        "num_os":             "8243",
        "contratante":        "LUCIA TSUYAKO ARASHIRO",
        "motorista_matricula": 2404,
        "motorista_nome":     "Carlos Luiz da Silva",
        "carro":              "0001283",
        "prefixo":            1283,
        "destino":            "SAO PAULO",
        "data_saida":         date(2026, 7, 30),
        "data_retorno":       date(2026, 7, 30),
        "hora_saida":         "17:30",
        "hora_retorno":       "23:30",
        "saida_garagem_km":   559882, "saida_garagem_hora":   hr(16, 30),
        "chegada_cliente_km": 559889, "chegada_cliente_hora": hr(16, 45),
        "saida_cliente_km":   559889, "saida_cliente_hora":   hr(17, 40),
        "chegada_destino_km": 559916, "chegada_destino_hora": hr(19, 5),
        "saida_destino_km":   559916, "saida_destino_hora":   hr(22, 40),
        "chegada_origem_km":  559937, "chegada_origem_hora":  hr(23, 20),
        "chegada_garagem_km": 559944, "chegada_garagem_hora": hr(23, 40),
        "comissao":  145.00,
        "valor_viagem": 1450.00,
    },
    {
        # OS 7667 · 31/07/26 · Carlos Luiz · Vargem Grande Paulista
        "num_os":             "7667",
        "contratante":        "MAURICIO FIGUEIREDO DE SOUSA",
        "motorista_matricula": 2404,
        "motorista_nome":     "Carlos Luiz da Silva",
        "carro":              "1283",
        "prefixo":            1283,
        "destino":            "VARGEM GRANDE PAULISTA",
        "data_saida":         date(2026, 7, 31),
        "data_retorno":       date(2026, 7, 31),
        "hora_saida":         "06:15",
        "hora_retorno":       "17:00",
        "saida_garagem_km":   559971, "saida_garagem_hora":   hr(5, 30),
        "chegada_cliente_km": 559985, "chegada_cliente_hora": hr(6, 0),
        "saida_cliente_km":   559985, "saida_cliente_hora":   hr(7, 0),
        "chegada_destino_km": 560040, "chegada_destino_hora": hr(8, 40),
        "saida_destino_km":   560040, "saida_destino_hora":   hr(16, 0),
        "chegada_origem_km":  560090, "chegada_origem_hora":  hr(18, 40),
        "chegada_garagem_km": 560104, "chegada_garagem_hora": hr(19, 20),
        "comissao":  190.00,
        "valor_viagem": 1900.00,
    },
    {
        # OS 7667 · 31/07/26 · Sandão Eduardo · Vargem Grande Paulista
        "num_os":             "7667",
        "contratante":        "MAURICIO FIGUEIREDO DE SOUSA",
        "motorista_matricula": 3049,
        "motorista_nome":     "Sandão Eduardo",
        "carro":              "0001433",
        "prefixo":            1433,
        "destino":            "VARGEM GRANDE PAULISTA",
        "data_saida":         date(2026, 7, 31),
        "data_retorno":       date(2026, 7, 31),
        "hora_saida":         "06:15",
        "hora_retorno":       "17:00",
        "saida_garagem_km":   531954, "saida_garagem_hora":   hr(5, 25),
        "chegada_cliente_km": 531981, "chegada_cliente_hora": hr(6, 30),
        "saida_cliente_km":   531981, "saida_cliente_hora":   hr(7, 30),
        "chegada_destino_km": 532041, "chegada_destino_hora": hr(8, 43),
        "saida_destino_km":   532091, "saida_destino_hora":   hr(16, 10),
        "chegada_origem_km":  532098, "chegada_origem_hora":  hr(18, 39),
        "chegada_garagem_km": 532120, "chegada_garagem_hora": hr(19, 50),
        "comissao":  190.00,
        "valor_viagem": 1900.00,
    },
]

# ── Executa criação ───────────────────────────────────────────────────────────
def main():
    for label, path in ARQUIVOS.items():
        if not path.exists():
            print(f"[ERRO] Arquivo não encontrado: {path}")

    # Km Turismo
    path = ARQUIVOS["km"]
    if path.exists():
        print(f"Abrindo {path.name}...")
        wb = load_workbook(str(path))
        criar_teste_km(wb, OS_EXEMPLOS)
        wb.save(str(path))
        print("  OK - Aba TESTE criada em Km Turismo")

    # Análise de Resultados
    path = ARQUIVOS["analise"]
    if path.exists():
        print(f"Abrindo {path.name}...")
        wb = load_workbook(str(path))
        criar_teste_analise(wb, OS_EXEMPLOS)
        wb.save(str(path))
        print("  OK - Aba TESTE criada em Analise de Resultados")

    # Comissão dos Motoristas
    path = ARQUIVOS["comissao"]
    if path.exists():
        print(f"Abrindo {path.name}...")
        wb = load_workbook(str(path))
        criar_teste_comissao(wb, OS_EXEMPLOS)
        wb.save(str(path))
        print("  OK - Aba TESTE criada em Comissao dos Motoristas")

    print("\nPronto! Abra os arquivos Excel e confira a aba TESTE em cada um.")

if __name__ == "__main__":
    main()
