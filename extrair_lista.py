"""
Extrator de Listas de Passageiros
----------------------------------
Lê imagens de listas de passageiros (fotografias) e extrai nomes
e números de documento para uma planilha XLSX editável.

Uso:
    python extrair_lista.py                         # processa todas as imagens na pasta
    python extrair_lista.py imagem1.jpg imagem2.jpg # processa imagens específicas
    python extrair_lista.py --saida resultado.xlsx  # define nome do arquivo de saída
"""

import argparse
import base64
import json
import os
import sys
from pathlib import Path

# Garante UTF-8 no terminal Windows
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Extensões de imagem suportadas
EXTENSOES_IMAGEM = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

MEDIA_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


def extrair_passageiros(client: Groq, caminho_imagem: Path) -> list[dict]:
    """
    Envia a imagem para o Groq e extrai a lista de passageiros.
    Retorna uma lista de dicionários com 'nome' e 'documento'.
    """
    print(f"  Analisando: {caminho_imagem.name}...")

    with open(caminho_imagem, "rb") as f:
        imagem_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
    media_type = MEDIA_TYPES[caminho_imagem.suffix.lower()]

    prompt = """Analise esta imagem de uma lista de passageiros e extraia TODOS os dados visíveis.

Para cada passageiro encontrado, retorne um objeto JSON com:
- "nome": nome completo do passageiro
- "documento": número do documento (RG, CPF, passaporte ou qualquer número de identificação presente)
- "observacao": qualquer informação adicional relevante (opcional, deixe vazio se não houver)

Se um campo não estiver visível ou legível, use null.

Responda SOMENTE com um array JSON válido, sem texto adicional, sem markdown, sem explicações.
Exemplo: [{"nome": "João Silva", "documento": "12.345.678-9", "observacao": ""}, ...]

Se a imagem não contiver lista de passageiros ou não for legível, retorne: []"""

    response = client.chat.completions.create(
        model="meta-llama/llama-4-scout-17b-16e-instruct",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{media_type};base64,{imagem_b64}"},
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
        max_tokens=4096,
        temperature=0,
    )
    texto = response.choices[0].message.content.strip()

    if not texto:
        print(f"  ⚠ Nenhuma resposta de texto para {caminho_imagem.name}")
        return []

    # Remove possível formatação markdown caso o modelo a inclua
    if texto.startswith("```"):
        linhas = texto.split("\n")
        texto = "\n".join(linhas[1:-1] if linhas[-1].strip() == "```" else linhas[1:])

    try:
        passageiros = json.loads(texto)
        if not isinstance(passageiros, list):
            print(f"  ⚠ Resposta inesperada para {caminho_imagem.name}: não é uma lista")
            return []
        print(f"  ✓ {len(passageiros)} passageiro(s) encontrado(s)")
        return passageiros
    except json.JSONDecodeError as e:
        print(f"  ⚠ Erro ao interpretar JSON para {caminho_imagem.name}: {e}")
        print(f"    Resposta recebida: {texto[:200]}...")
        return []


def criar_planilha(passageiros: list[dict], caminho_saida: Path, nome_origem: str) -> None:
    """Cria o arquivo XLSX com os dados de uma única imagem."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Passageiros"

    cor_cabecalho = "2B5EA7"
    fonte_cabecalho = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    preenchimento_cabecalho = PatternFill(fill_type="solid", fgColor=cor_cabecalho)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    cor_linha_par = "DCE6F1"
    preenchimento_par = PatternFill(fill_type="solid", fgColor=cor_linha_par)

    cabecalhos = ["✓", "#", "Nome Completo", "Documento", "Observação"]
    larguras =   [5,   6,   40,              20,           30]

    for col, (cabecalho, largura) in enumerate(zip(cabecalhos, larguras), start=1):
        celula = ws.cell(row=1, column=col, value=cabecalho)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = alinhamento_centro
        ws.column_dimensions[celula.column_letter].width = largura

    ws.row_dimensions[1].height = 22

    for i, p in enumerate(passageiros, start=1):
        linha = i + 1
        preenchimento = preenchimento_par if linha % 2 == 0 else None

        # Coluna ✓ — checkbox via validação de lista (☐ / ☑)
        cel_check = ws.cell(row=linha, column=1, value="☐")
        cel_check.alignment = Alignment(horizontal="center", vertical="center")
        cel_check.font = Font(name="Calibri", size=14)

        valores = [i, p.get("nome") or "", p.get("documento") or "", p.get("observacao") or ""]
        for col, valor in enumerate(valores, start=2):
            celula = ws.cell(row=linha, column=col, value=valor)
            celula.alignment = Alignment(vertical="center", wrap_text=True)

        if preenchimento:
            for col in range(1, len(cabecalhos) + 1):
                ws.cell(row=linha, column=col).fill = preenchimento

        ws.row_dimensions[linha].height = 18

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"B1:{ws.cell(row=1, column=len(cabecalhos)).coordinate}"
    wb.save(caminho_saida)


def listar_imagens(pasta: Path) -> list[Path]:
    """Lista todas as imagens suportadas na pasta."""
    return sorted(
        p for p in pasta.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSOES_IMAGEM
    )


def main():
    parser = argparse.ArgumentParser(
        description="Extrai listas de passageiros de imagens para XLSX",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("imagens", nargs="*", help="Arquivos de imagem a processar")
    parser.add_argument("--pasta", default=".", help="Pasta onde procurar imagens")
    args = parser.parse_args()

    pasta = Path(args.pasta)

    if args.imagens:
        arquivos = [Path(i) for i in args.imagens]
        nao_encontrados = [f for f in arquivos if not f.exists()]
        if nao_encontrados:
            for f in nao_encontrados:
                print(f"Arquivo não encontrado: {f}")
            sys.exit(1)
    else:
        arquivos = listar_imagens(pasta)
        if not arquivos:
            print(f"Nenhuma imagem encontrada em: {pasta.resolve()}")
            sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  EXTRATOR DE LISTAS DE PASSAGEIROS")
    print(f"{'='*60}")
    print(f"  Imagens a processar: {len(arquivos)}")
    print(f"{'='*60}\n")

    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        print("ERRO: Variável de ambiente GROQ_API_KEY não definida.")
        print("Obtenha sua chave gratuita em: https://console.groq.com")
        print("Configure com: set GROQ_API_KEY=sua_chave_aqui")
        sys.exit(1)

    client = Groq(api_key=api_key)
    total_passageiros = 0
    planilhas_geradas = []

    for i, arquivo in enumerate(arquivos, start=1):
        print(f"[{i}/{len(arquivos)}] {arquivo.name}")
        passageiros = extrair_passageiros(client, arquivo)

        if not passageiros:
            print(f"  ⚠ Nenhum passageiro extraído — planilha não gerada.\n")
            continue

        # Nome do XLSX = nome da imagem (sem extensão) + .xlsx
        caminho_saida = arquivo.parent / (arquivo.stem + ".xlsx")
        criar_planilha(passageiros, caminho_saida, arquivo.name)
        print(f"  ✓ Planilha salva: {caminho_saida.name}\n")
        planilhas_geradas.append(caminho_saida)
        total_passageiros += len(passageiros)

    print(f"{'='*60}")
    print(f"  Imagens processadas: {len(arquivos)}")
    print(f"  Planilhas geradas:   {len(planilhas_geradas)}")
    print(f"  Total de passageiros: {total_passageiros}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
